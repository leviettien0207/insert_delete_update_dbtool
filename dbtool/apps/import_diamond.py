import openpyxl
import pandas
from rest_framework.views import APIView
from rest_framework import permissions
from rest_framework.response import Response
from .constant import *
from .Item import Item
from .models import diamond
import os


class Import(APIView):
    permission_classes = [permissions.AllowAny]

    errors = []
    row_errors = []
    success_Item = []

    def post(self, request, filepath):
        self.reset_list()  # todo kiểm tra liệu có cần reset không?
        path_file = os.path.join(PATH_FILE, filepath)
        excel_handle = self.open_excel(path_file)
        sheet_handle = self.open_sheet(excel_handle, 'diamonds')
        if len(self.errors) != 0:
            FAIL_OVERALL["ERRORS"] = FAIL_OVERALL["ERRORS"].format(self.errors)
            return Response(FAIL_OVERALL)

        # take data from sheet
        df = pandas.DataFrame(sheet_handle.values)

        # new_header = df.iloc[0] #grab the first row for the header
        # df = df[1:] #take the data less the header row
        # df.columns = new_header #set the header row as the df header
        # df.drop_duplicates(subset=['id'], inplace=True)
        # excel_handle.save(PATH_FILE)
        count = 0
        for idx, row_data in df.iterrows():
            self.row_errors.clear()
            if idx == 0:
                continue  # ignore first row
            row_number = idx + 1

            self.writeCell(sheet_handle, row_number, DIAMOND_COLUMNS, '')
            if self.isCommand(row_data):  # check exist command
                self.execute_row(sheet_handle, row_number, row_data)  # --> add db (if possible) write result
            else:
                self.writeCell(sheet_handle, row_number, DIAMOND_COLUMNS, MSG_WRONG_COMMAND.format(row_data[0]))
            if count == 200 :
                diamond.objects.bulk_create(self.success_Item)
                excel_handle.save(path_file)
                self.success_Item.clear()
                count = 0
            print('processing....row{} ok'.format(row_number))
        diamond.objects.bulk_create(self.success_Item)
        excel_handle.save(path_file)
        

        return Response(SUCCESS_200)

    def open_excel(self, path_file):
        try:
            return openpyxl.load_workbook(path_file)  # todo catch errors when openpyxl does not supper extensions
        except FileNotFoundError:
            self.errors.append(FAIL_OPEN_EXCEL.format(path_file))

    def open_sheet(self, work_book, work_sheet_name):
        if work_sheet_name in work_book.sheetnames:
            return work_book[work_sheet_name]
        else:
            self.errors.append(FAIL_OPEN_SHEET.format(work_sheet_name))

    def reset_list(self):
        self.errors.clear()
        self.success_Item.clear()

    def isCommand(self, row_data):  # check existence db_command
        if row_data[0] is not None and row_data[0].upper() in DB_COMMAND_RANGE:
            return True
        else:
            return False

    def execute_row(self, ws, row_number, row_data):
        self.row_errors.clear()  # todo whether write all errors or not
        if row_data[0].upper() == 'INSERT':
            self.exec_insert(ws, row_number, row_data)
        elif row_data[0].upper() == 'DELETE':
            self.exec_delete(ws, row_number, row_data)
        elif row_data[0].upper() == 'UPDATE':
            self.exec_update(ws, row_number, row_data)

    @staticmethod
    def writeCell(ws, row, col, msg):
        ws.cell(row=row, column=col, value=msg)

    def exec_insert(self, ws, row_number, row_data):
        if self.validate_id(row_data):
            self.writeCell(ws, row_number, DIAMOND_COLUMNS, MSG_MISSING_INFO.format(row_data[1]))

        find_record = diamond.objects.filter(id=row_data[1])
        if len(find_record) == 0:
            item = Item(row_data)
            item.validate_all()
            if len(item.data_errors) != 0:
                self.writeCell(ws, row_number, DIAMOND_COLUMNS, str(item.data_errors))
            else:
                toDiamond = item.convertDiamond()
                self.success_Item.append(toDiamond)
                self.writeCell(ws, row_number, DIAMOND_COLUMNS, SUCCESS_200)
        else:
            self.writeCell(ws, row_number, DIAMOND_COLUMNS, EXIST_ID.format(row_data[1]))

    def exec_delete(self, ws, row_number, row_data):
        if self.validate_id(row_data):
            self.writeCell(ws, row_number, DIAMOND_COLUMNS, MSG_MISSING_INFO.format(row_data[1]))

        find_record = diamond.objects.filter(id=row_data[1])
        if len(find_record) == 0:
            self.writeCell(ws, row_number, DIAMOND_COLUMNS, FAIL_RECORD.format(row_data[1]))
        else:
            find_record.delete()
            self.writeCell(ws, row_number, DIAMOND_COLUMNS, SUCCESS_200)

    def exec_update(self, ws, row_number, row_data):
        if self.validate_id(row_data):
            self.writeCell(ws, row_number, DIAMOND_COLUMNS, MSG_MISSING_INFO.format(row_data[1]))

        find_record = diamond.objects.filter(id=row_data[1])
        if len(find_record) == 0:
            self.writeCell(ws, row_number, DIAMOND_COLUMNS, FAIL_RECORD.format(row_data[1]))
        else:
            self.exec_delete(ws, row_number, row_data)
            self.exec_insert(ws, row_number, row_data)
            self.writeCell(ws, row_number, DIAMOND_COLUMNS, SUCCESS_200)

    @staticmethod
    def validate_id(row_data):
        return True if row_data[1] is not None else False

